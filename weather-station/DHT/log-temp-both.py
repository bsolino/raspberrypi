"""
Record temperature and humidity and save it to disk

Initial DHT code from https://tutorials-raspberrypi.com/raspberry-pi-measure-humidity-temperature-dht11-dht22/
"""

import time
import board
import adafruit_dht
from datetime import datetime
from openpyxl import Workbook, load_workbook
from pathlib import Path

# Initial the dht device, with data pin connected to:
# dhtDevice = adafruit_dht.DHT22(board.D4)

# you can pass DHT22 use_pulseio=False if you wouldn't like to use pulseio.
# This may be necessary on a Linux single board computer like the Raspberry Pi,
# but it will not work in CircuitPython.


# Configuration

DEBUG = True
PIN_11 = board.D4
PIN_22 = board.D17
OUTPUT_FOLDER = Path(Path.home(), "weather")
MEASURE_WAIT = 5 * 60 # Time between measures (in seconds)
RETRY_WAIT = 2 # Time between retries (in seconds)
N_MEASURES = 5

# Functions
def take_measurement(dhtDevice):
    temperature = dhtDevice.temperature
    humidity = dhtDevice.humidity
    
    if temperature is None or humidity is None:
        raise RuntimeError(f"DHT Returned a None value: T: {temperature}, H: {humidity}")
    else:
        return temperature, humidity

def record_weather(dhtDevice, name, measurements):
    temperature, humidity = take_measurement(dhtDevice)

    # Print and record:

    now = datetime.now()
    msg = f"{name}: {now.date()} {now.time()}"
    msg += f"\tTemp: {temperature:.1f} C\tHumid: {humidity} %"
    
    if len(measurements) == 0: # Discard first measure
        msg += " Discarded"
    row = (name, now.date(), now.time(), temperature, humidity)
    measurements.append(row)
    print(msg)

# Initialization

OUTPUT_FOLDER.mkdir(parents = True, exist_ok=True)
WB_NAME = "{date}-weather.xlsx"
SHEET = "Sheet"

dht11 = adafruit_dht.DHT11(PIN_11)
dht22 = adafruit_dht.DHT22(PIN_22)

#dhtDevice = adafruit_dht.DHT11(PIN, use_pulseio=False)

wb_date = datetime.now().date()
wb_name = WB_NAME.format(date = wb_date)
wb_path = Path(OUTPUT_FOLDER, wb_name)
try:
    wb = load_workbook(wb_path)
except FileNotFoundError as error:
    print(error)
    print(f"Creating new file {wb_path}")
    wb = Workbook()
sheet = wb[SHEET]

measurements11 = list()
measurements22 = list()
do_measure11 = do_measure22 = True
while True:
#    date = datetime.now().date()
#    if date > wb_date(): # New wb per day

    try:
        do_measure11 = do_measure11 and len(measurements11) < N_MEASURES+1
        do_measure22 = do_measure22 and len(measurements22) < N_MEASURES+1
        if do_measure11:
            name = "DHT11"
            do_measure11 = False
            record_weather(dht11, name, measurements11)
        if do_measure22:
            name = "DHT22"
            do_measure22 = False
            record_weather(dht22, name, measurements22)
        
        print(f"{len(measurements11)}, {len(measurements22)}")
        if len(measurements11) >= N_MEASURES+1 \
                and len(measurements22) >= N_MEASURES+1:
            for row in (*measurements11[1:], *measurements22[1:]):
                sheet.append(row)
            wb.save(wb_path)
            measurements11.clear()
            measurements22.clear()
            wait11 = wait22 = False
            print("Sleep")
            time.sleep(MEASURE_WAIT)
        elif not do_measure11 or not do_measure22:
            do_measure11 = do_measure22 = True
            if DEBUG:
                print(f"Waiting for new measure: {RETRY_WAIT}s")
            time.sleep(RETRY_WAIT)
        else:
            print(f"Not waiting: {not do_measure11 or not do_measure22}")

    except KeyboardInterrupt as error:
        print("Exiting...")
        dht11.exit()
        dht22.exit()
        wb.save(wb_path) # In case it's killed while saving
        raise error

    except RuntimeError as error:
        # Errors happen fairly often, DHT's are hard to read, just keep going
        if DEBUG:
            print(f"{name}: {error.args[0]}")
    except Exception as error:
        print("Exiting...")
        dht11.exit()
        dht22.exit()
        wb.save(wb_path) # In case it's killed while saving
        raise error


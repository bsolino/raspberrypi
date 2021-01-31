"""
Record temperature and humidity and save it to disk

Initial DHT code from https://tutorials-raspberrypi.com/raspberry-pi-measure-humidity-temperature-dht11-dht22/
"""

import time
import board
import adafruit_dht
from datetime import datetime
from openpyxl import Workbook, load_workbook
from pathlib import Path

# Initial the dht device, with data pin connected to:
# dhtDevice = adafruit_dht.DHT22(board.D4)

# you can pass DHT22 use_pulseio=False if you wouldn't like to use pulseio.
# This may be necessary on a Linux single board computer like the Raspberry Pi,
# but it will not work in CircuitPython.


# Configuration

PIN = board.D4
OUTPUT_FOLDER = Path(Path.home(), "weather")
MEASURE_WAIT = 10 * 60 # Time between measures (in seconds)
RETRY_WAIT = 2. # Time between retries (in seconds)
N_MEASURES = 5

# Initialization

OUTPUT_FOLDER.mkdir(parents = True, exist_ok=True)
WB_NAME = "weather.xlsx"
SHEET = "Sheet"
dhtDevice = adafruit_dht.DHT11(PIN)
#dhtDevice = adafruit_dht.DHT11(PIN, use_pulseio=False)

#wb_date = datetime.now().date()

wb_path = Path(OUTPUT_FOLDER, WB_NAME)
try:
    wb = load_workbook(wb_path)
except FileNotFoundError as error:
    print(error)
    print("Creating new file")
    wb = Workbook()
sheet = wb[SHEET]

i_measure = 0
while True:
#    date = datetime.now().date()
#    if date > wb_date(): # New wb per day

    try:
        # Print the values to the serial port
        temperature = dhtDevice.temperature
        humidity = dhtDevice.humidity
        
        if temperature is None or humidity is None:
            raise RuntimeError(f"DHT Returned a None value: T: {temperature}, H: {humidity}")

        now = datetime.now()
        print(f"{now.date()} {now.time()}\tTemp: {temperature:.1f} C\tHumidity: {humidity} %")

        i_measure += 1
        row = (now.date(), now.time(), temperature, humidity)
        sheet.append(row)


    except RuntimeError as error:
        # Errors happen fairly often, DHT's are hard to read, just keep going
#        print(error.args[0])
        time.sleep(RETRY_WAIT)
        continue
    except Exception as error:
        dhtDevice.exit()
        wb.save(wb_path) # In case it's killed while saving
        raise error

    if i_measure == N_MEASURES:
        wb.save(wb_path)
        i_measure = 0
        print("Sleep")
        time.sleep(MEASURE_WAIT)
    else:
        time.sleep(RETRY_WAIT)

